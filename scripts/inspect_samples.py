"""네이버 파트너센터 xlsx 샘플 구조 분석. 결과를 samples/_inspect.txt에 UTF-8로 저장."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

SAMPLES_DIR = Path(__file__).resolve().parent.parent / "samples"
OUT_PATH = SAMPLES_DIR / "_inspect.txt"


def inspect(path: Path, out) -> None:
    out.write("\n" + "=" * 80 + "\n")
    out.write(f"FILE: {path.name}\n")
    out.write("=" * 80 + "\n")

    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        out.write(f"\n[Sheet] {sheet_name}  (rows={ws.max_row}, cols={ws.max_column})\n")

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            out.write("  (empty)\n")
            continue

        head = rows[:15]
        tail = rows[-3:] if len(rows) > 18 else []

        for i, row in enumerate(head, 1):
            out.write(f"  {i:3d}: {row}\n")
        if tail:
            out.write(f"  ... ({len(rows) - 15} more rows) ...\n")
            for i, row in enumerate(tail, len(rows) - 2):
                out.write(f"  {i:3d}: {row}\n")


def main() -> None:
    with OUT_PATH.open("w", encoding="utf-8") as out:
        for path in sorted(SAMPLES_DIR.glob("*.xlsx")):
            inspect(path, out)
    print(f"Saved to {OUT_PATH}")


if __name__ == "__main__":
    main()
