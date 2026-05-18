"""네이버 파트너센터 xlsx 파서.

각 메뉴별 xlsx 파일은 공통적으로 1~7행이 메타데이터(서비스명/데이터명/기간 등),
빈 행을 사이에 두고 실제 데이터가 시작된다. 메타데이터에서 data_date를 뽑고,
데이터 영역을 메뉴별 dataclass로 파싱한다.
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime, timezone, timedelta
from pathlib import Path

from openpyxl import load_workbook

KST = timezone(timedelta(hours=9))

_DATE_RE = re.compile(r"(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일")
_DATETIME_RE = re.compile(r"(\d{4})\.(\d{1,2})\.(\d{1,2})\.?\s*(\d{1,2}):(\d{1,2})")


@dataclass
class Metadata:
    """xlsx 헤더 7행에서 추출한 메타정보."""
    data_name: str          # '기사 조회수 순위' / '시간대별 조회수' / '유입분석' / '유입키워드'
    data_date: date          # 데이터 기간 날짜
    downloaded_at: datetime  # 다운로드 시각 (KST)


@dataclass
class ArticlePvRow:
    rank: int
    title: str
    reporter_name: str | None
    article_published_at: datetime  # KST
    pv: int


@dataclass
class HourlyPvRow:
    hour: int
    pv: int


@dataclass
class TrafficSourceRow:
    source_category: str
    category_ratio: float
    source_detail_url: str | None
    detail_ratio: float


@dataclass
class SearchKeywordRow:
    rank: int
    keyword: str
    clicks: int
    ratio: float


def _extract_date(s: str) -> date:
    m = _DATE_RE.search(s)
    if not m:
        raise ValueError(f"날짜 추출 실패: {s}")
    return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))


def _extract_datetime_kst(s: str) -> datetime:
    """'2026.05.17. 05:02' 형식 → KST aware datetime."""
    m = _DATETIME_RE.search(s)
    if not m:
        raise ValueError(f"datetime 추출 실패: {s}")
    y, mo, d, h, mi = map(int, m.groups())
    return datetime(y, mo, d, h, mi, tzinfo=KST)


def _parse_metadata(rows: list[tuple]) -> Metadata:
    meta_map: dict[str, str] = {}
    for row in rows[:10]:
        if not row or row[0] is None:
            continue
        key = str(row[0]).strip()
        val = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        meta_map[key] = val

    data_name = meta_map.get("데이터명", "")
    data_date = _extract_date(meta_map.get("데이터 기간", ""))

    # '2026년 05월 18일 21시 41분 06초' → datetime
    raw = meta_map.get("다운로드 날짜", "")
    m = re.search(r"(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일\s*(\d{1,2})시\s*(\d{1,2})분\s*(\d{1,2})초", raw)
    if m:
        y, mo, d, h, mi, se = map(int, m.groups())
        downloaded_at = datetime(y, mo, d, h, mi, se, tzinfo=KST)
    else:
        downloaded_at = datetime.now(KST)

    return Metadata(data_name=data_name, data_date=data_date, downloaded_at=downloaded_at)


def _data_rows(rows: list[tuple]) -> list[tuple]:
    """헤더 7행 + 빈행 + 컬럼명 1행 이후의 실데이터만 반환. 마지막 빈행 제거."""
    # 빈 행을 찾아 그 다음 1행을 컬럼명, 그 다음부터 데이터로 간주
    for i, row in enumerate(rows):
        if row[0] is None and i > 0:
            # i+1행이 컬럼명, i+2부터 데이터
            return [r for r in rows[i + 2:] if r and r[0] is not None]
    return []


def parse_article_pv(path: Path) -> tuple[Metadata, list[ArticlePvRow]]:
    rows = _load_rows(path)
    meta = _parse_metadata(rows)
    items: list[ArticlePvRow] = []
    for r in _data_rows(rows):
        rank = int(r[0])
        title = str(r[1]).strip()
        reporter = str(r[2]).strip() if r[2] else None
        published_at = _extract_datetime_kst(str(r[3]))
        pv = int(r[4])
        items.append(ArticlePvRow(rank, title, reporter, published_at, pv))
    return meta, items


def parse_hourly_pv(path: Path) -> tuple[Metadata, list[HourlyPvRow]]:
    rows = _load_rows(path)
    meta = _parse_metadata(rows)
    items: list[HourlyPvRow] = []
    for r in _data_rows(rows):
        hour = int(r[0])
        pv = int(r[1])
        items.append(HourlyPvRow(hour, pv))
    return meta, items


def parse_traffic_source(path: Path) -> tuple[Metadata, list[TrafficSourceRow]]:
    rows = _load_rows(path)
    meta = _parse_metadata(rows)
    items: list[TrafficSourceRow] = []
    for r in _data_rows(rows):
        source_category = str(r[0]).strip()
        category_ratio = float(r[1])
        source_detail_url = str(r[2]).strip() if r[2] else None
        detail_ratio = float(r[3])
        items.append(
            TrafficSourceRow(source_category, category_ratio, source_detail_url, detail_ratio)
        )
    return meta, items


def parse_search_keyword(path: Path) -> tuple[Metadata, list[SearchKeywordRow]]:
    rows = _load_rows(path)
    meta = _parse_metadata(rows)
    items: list[SearchKeywordRow] = []
    for r in _data_rows(rows):
        rank = int(r[0])
        keyword = str(r[1]).strip()
        clicks = int(r[2])
        ratio = float(r[3])
        items.append(SearchKeywordRow(rank, keyword, clicks, ratio))
    return meta, items


def _load_rows(path: Path) -> list[tuple]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    return list(ws.iter_rows(values_only=True))
